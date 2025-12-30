import openpyxl
book = openpyxl.load_workbook("C:\\Users\\tpss\\PycharmProjects\\TestExcelFile\\PythonDemo1.xlsx")

sheet = book.active
cell = sheet.cell(row=1, column = 2)
print(cell.value)

sheet.cell(row=2, column = 2).value = "Vishal QA"

print(sheet.cell(row = 2, column =2).value)